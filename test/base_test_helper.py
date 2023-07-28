import os
from pathlib import PurePosixPath, Path

import allure
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from seleniumbase import BaseCase

BaseCase.main(__name__, __file__)

STEPS_RTM_DIR = PurePosixPath.joinpath(Path(__file__).parent.parent).joinpath('steps_rtm')


class BaseTestHelper(BaseCase):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.remove_text_from_step_rtm()

    def setUp_(self):
        super().setUp()

    def tearDown_(self):
        self.save_teardown_screenshot()
        self.get_screenshot_from_allure()
        self.evidence_name = self.get_file_name("evidence_")
        self.get_save_screenshot_name(self.evidence_name)
        super().tearDown()

    @allure.step("Open browser url")
    def open_browser_url(self, url: str):
        self.get_stdout_from_allure(f'OPEN URL: {url}')
        return self.open(url)

    @allure.step("Assert text element")
    def assertion_text_element(self, text: str, element):
        text_compare = self.get_text(element)
        self.get_stdout_from_allure(f"COMPARE TEXT: {text} VS: {text_compare}")
        return self.assert_exact_text(text, element)

    @allure.step("Assert element")
    def assert_element_(self, element):
        self.get_stdout_from_allure(f"ASSERT ELEMENT: {element}")
        return self.assert_element(element)

    @allure.step("Writing text on element: {element}")
    def type_(self, element, text: str):
        self.get_stdout_from_allure(f"WRITING TEXT: {text} IN ELEMENT: {element}")
        return self.type(element, text)

    @allure.step("Click on element")
    def click_on_element(self, element):
        self.get_stdout_from_allure(f"CLICK ON ELEMENT: {element}")
        return self.click(element)

    def sleep_(self, seconds):
        return self.sleep(seconds)

    @allure.step("Switch windows handler")
    def switch_to_windows_handle(self, name):
        self.get_stdout_from_allure(f"SWITCH TO WINDOWS: {name}")
        return self.switch_to_window(name)

    @allure.step("Search text in page")
    def search_text_in_page(self, page_number, search_text):
        self.switch_to_window(page_number)
        context_page = self.get_page_source()
        if search_text in context_page:
            self.get_stdout_from_allure(f'The text "{search_text}" is in a page')
        else:
            self.get_stdout_from_allure(f'Test failed! The text "{search_text}" was not found')
            assert False

    @allure.step("Take a screnshot")
    def save_screenshot_to_dir(self, name, folder, selector):
        self.get_stdout_from_allure(f'TAKE SCREENSHOT WITH NAME: {name} IN FOLDER: {folder}')
        return self.save_screenshot(name, folder, selector)

    @allure.step("Switch to tab")
    def switch_to_tab_in_browser(self, tab):
        self.get_stdout_from_allure(f"SWITCH TO TAB: {tab}")
        return self.switch_to_tab(tab)

    @allure.step("Switch to alert")
    def compare_test_from_alert(self, time, text_to_compare):
        alert = self.wait_for_and_switch_to_alert(time)
        self.sleep_(time)
        text_alert = alert.text
        self.get_stdout_from_allure(f"CAPTURING TEXT FROM ALERT: {text_alert}")
        assert text_alert == text_to_compare
        self.wait_for_and_accept_alert(time)

    @allure.step("Find elements to list")
    def find_elements_to_list(self, elements):
        self.get_stdout_from_allure(f"ELEMENTS TO LIST: {elements}")
        return self.find_elements(elements)

    def iter_lists_for_course(self, list_one, list_two, value):
        for course, price in zip(list_one, list_two):
            c = course.text
            p = int(price.text)
            if p == value:
                self.get_stdout_from_allure(f"COURSE: {c}")

    @allure.step("Switch to frame")
    def switch_to_frame_(self, element):
        self.get_stdout_from_allure(f"SWITCH TO FRAME: {element}")
        return self.switch_to_frame(element)





    def get_screenshot_from_allure(self):
        img = allure.attach(body=self.driver.get_screenshot_as_png(), name="screenshot",
                            attachment_type=allure.attachment_type.PNG)
        return img

    def get_evidence(self):
        img = self.get_screenshot_from_allure()
        return img

    def get_stdout_from_allure(self, print_texto):
        step_rtm = os.path.join(STEPS_RTM_DIR, "step_rtm.txt")
        print(print_texto)
        with open(step_rtm, "a") as archivo:
            archivo.write(print_texto + "\n")

    def remove_text_from_step_rtm(self):
        step_rtm = os.path.join(STEPS_RTM_DIR, "step_rtm.txt")
        with open(step_rtm, "w") as archivo:
            archivo.truncate()

    def get_steps_from_step_rtm(self):
        step_rtm = os.path.join(STEPS_RTM_DIR, "step_rtm.txt")
        try:
            with open(step_rtm, "r") as origin_file:
                content = origin_file.read()
                if content is not None:
                    return content
        except FileNotFoundError as f:
            print("FileNotFoundError: ", f)
        except Exception as e:
            print("Error:", e)


    def get_save_screenshot_name(self, name):
        self.save_screenshot(name, f"{STEPS_RTM_DIR}/evidence")

    def get_actual_result(self):
        if self.get_status_test() == 'passed':
            resp = "Same as expected result"
        else:
            resp = "Not equal to expected result"
        return resp

    def get_status_test(self):
        status_test = os.path.join(STEPS_RTM_DIR, "status_test.txt")
        with open(status_test, "r") as status:
            content = status.read()
            if content is not None:
                return content


    def set_rtm_with_data(self, data):
        path_excel = os.path.join(STEPS_RTM_DIR, "rtm_test_cases.xlsx")

        workbook = load_workbook(path_excel)
        worksheet = workbook["Test Cases"]
        self.write_titles(worksheet)

        test_case = TestCase(**data)

        for row in worksheet.iter_rows(max_row=40):
            cell_value = row[0].value
            cell_int_value = row[0].row
            if cell_value is None:
                self.write_object_in_a_row(test_case=test_case, cell_row=cell_int_value, worksheet=worksheet)
                break
        self.set_image_evidence(workbook=workbook)
        self.set_hipervinculo(workbook)

    def set_image_evidence(self, workbook):
        image_path = os.path.join(STEPS_RTM_DIR, "evidence", self.evidence_name)

        img = Image(image_path)
        img.width = 1024
        img.height = 648

        sheet_name = self.evidence_name
        if sheet_name in workbook.sheetnames:
            pass
        else:
            workbook.create_sheet(title=sheet_name)

        hoja_evidence = workbook[sheet_name]

        for row in hoja_evidence.iter_rows(max_row=1):
            for cell in row:
                cell_value = cell.value
                if cell_value is None:
                    cell_trazability = f"{cell.column_letter}{cell.row}"
                    hoja_evidence.add_image(img, cell_trazability)
                else:
                    pass


    def set_hipervinculo(self, workbook):
        path_excel = os.path.join(STEPS_RTM_DIR, "rtm_test_cases.xlsx")

        worksheet = workbook["Test Cases"]

        for row in worksheet.iter_rows(max_row=20):
            for cell in row:
                cell_value = cell.value
                if cell_value == "img":
                    celda_str = f"{cell.column_letter}{cell.row}"
                    hipervinculo = f"file://{path_excel}#'{self.evidence_name}'!{celda_str}"
                    worksheet[celda_str].hyperlink = hipervinculo
                    worksheet[celda_str].value = "Ver imagen"

        workbook.save(path_excel)


    def generate_file_name(self, name, cont):
        return f"{name}{cont}.png"

    def get_file_name(self, file_name):
        path_evidence = os.path.join(STEPS_RTM_DIR, "evidence")
        contador = 1
        new_name_file = f"{file_name}{contador}.png"
        while os.path.exists(os.path.join(path_evidence, new_name_file)):
            if contador == 1:
                new_name_file = self.generate_file_name(file_name, contador + 1)
            else:
                new_name_file = self.generate_file_name(file_name, contador)
            contador += 1

        return new_name_file


    def write_titles(self, worksheet):
        worksheet.cell(row=1, column=1, value="Test Case ID")
        worksheet.cell(row=1, column=2, value="Developer")
        worksheet.cell(row=1, column=3, value="Test group")
        worksheet.cell(row=1, column=4, value="Test case name")
        worksheet.cell(row=1, column=5, value="Priority/Severity")
        worksheet.cell(row=1, column=6, value="Pre-conditions")
        worksheet.cell(row=1, column=7, value="Test data")
        worksheet.cell(row=1, column=8, value="Test Steps")
        worksheet.cell(row=1, column=9, value="Expected Result")
        worksheet.cell(row=1, column=10, value="Tester")
        worksheet.cell(row=1, column=11, value="Actual result")
        worksheet.cell(row=1, column=12, value="Status")
        worksheet.cell(row=1, column=13, value="Comments")
        worksheet.cell(row=1, column=14, value="Test Evidence")

    def write_object_in_a_row(self, test_case, cell_row: int, worksheet):
        worksheet.cell(row=cell_row, column=1, value=test_case.test_case_id)
        worksheet.cell(row=cell_row, column=2, value=test_case.developer)
        worksheet.cell(row=cell_row, column=3, value=test_case.test_group)
        worksheet.cell(row=cell_row, column=4, value=test_case.test_case_name)
        worksheet.cell(row=cell_row, column=5, value=test_case.priority_severity)
        worksheet.cell(row=cell_row, column=6, value=test_case.pre_conditions)
        worksheet.cell(row=cell_row, column=7, value=test_case.test_data)
        worksheet.cell(row=cell_row, column=8, value=test_case.test_steps)
        worksheet.cell(row=cell_row, column=9, value=test_case.expected_result)
        worksheet.cell(row=cell_row, column=10, value=test_case.tester)
        worksheet.cell(row=cell_row, column=11, value=test_case.actual_result)
        worksheet.cell(row=cell_row, column=12, value=test_case.status)
        worksheet.cell(row=cell_row, column=13, value=test_case.comments)
        worksheet.cell(row=cell_row, column=14, value=test_case.test_evidence)


class TestCase:
    def __init__(self, test_case_id, developer, test_group, test_case_name, priority_severity,
                 pre_conditions, test_data, test_steps, expected_result, tester, actual_result,
                 status, comments, test_evidence):
        self.test_case_id = test_case_id
        self.developer = developer
        self.test_group = test_group
        self.test_case_name = test_case_name
        self.priority_severity = priority_severity
        self.pre_conditions = pre_conditions
        self.test_data = test_data
        self.test_steps = test_steps
        self.expected_result = expected_result
        self.tester = tester
        self.actual_result = actual_result
        self.status = status
        self.comments = comments
        self.test_evidence = test_evidence
